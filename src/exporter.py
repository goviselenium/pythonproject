import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.config import OUTPUT_DIR

def export_to_csv(jobs: List[Dict[str, Any]], filename: str = "job_matches_report.csv") -> Path:
    """Exports analyzed jobs to a CSV file."""
    output_path = OUTPUT_DIR / filename
    
    # Flatten JSON fields for tabular format
    flat_jobs = []
    for job in jobs:
        # Load JSON fields safely
        matched = _parse_json_field(job.get("matched_skills"))
        missing = _parse_json_field(job.get("missing_skills"))
        
        flat_jobs.append({
            "Score": job.get("score"),
            "Recommendation": job.get("recommendation"),
            "Title": job.get("title"),
            "Company": job.get("company"),
            "Location": job.get("location"),
            "URL": job.get("url"),
            "Matched Skills": ", ".join(matched) if isinstance(matched, list) else "",
            "Missing Skills": ", ".join(missing) if isinstance(missing, list) else "",
            "Explanation": job.get("explanation"),
            "Posted Date": job.get("posted_at"),
            "Scraped Date": job.get("scraped_at")
        })
        
    df = pd.DataFrame(flat_jobs)
    # Sort by score descending
    if not df.empty:
        df = df.sort_values(by="Score", ascending=False)
        
    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    return output_path

def export_to_excel(jobs: List[Dict[str, Any]], filename: str = "job_matches_report.xlsx") -> Path:
    """
    Exports analyzed jobs to a styled Excel sheet.
    Applies custom headers, colors, auto-fits columns, and aligns text.
    """
    output_path = OUTPUT_DIR / filename
    
    # Prepare flat list of jobs
    flat_jobs = []
    for job in jobs:
        matched = _parse_json_field(job.get("matched_skills"))
        missing = _parse_json_field(job.get("missing_skills"))
        
        flat_jobs.append({
            "Score": job.get("score") if job.get("score") is not None else 0,
            "Recommendation": job.get("recommendation") or "Pending",
            "Title": job.get("title") or "",
            "Company": job.get("company") or "",
            "Location": job.get("location") or "",
            "URL": job.get("url") or "",
            "Matched Skills": ", ".join(matched) if isinstance(matched, list) else "",
            "Missing Skills": ", ".join(missing) if isinstance(missing, list) else "",
            "Explanation": job.get("explanation") or "",
            "Posted Date": job.get("posted_at") or "",
        })
        
    df = pd.DataFrame(flat_jobs)
    if not df.empty:
        df = df.sort_values(by="Score", ascending=False)
    else:
        # Create empty template structure
        df = pd.DataFrame(columns=[
            "Score", "Recommendation", "Title", "Company", 
            "Location", "URL", "Matched Skills", "Missing Skills", 
            "Explanation", "Posted Date"
        ])

    # Write using pandas ExcelWriter
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Job Matches", index=False)
        
        # Get workbook and worksheet objects to style
        workbook = writer.book
        worksheet = writer.sheets["Job Matches"]
        
        # Define styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
        
        fill_apply = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
        font_apply = Font(name="Calibri", size=11, color="375623", bold=True)
        
        fill_customize = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow/Orange
        font_customize = Font(name="Calibri", size=11, color="7F6000", bold=True)
        
        fill_skip = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") # Soft Red
        font_skip = Font(name="Calibri", size=11, color="C65911")
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Style headers
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[1].height = 26
            
        # Style rows and conditional formatting on recommendation
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            rec_cell = worksheet.cell(row=row_num, column=2) # Recommendation is column 2
            score_cell = worksheet.cell(row=row_num, column=1) # Score is column 1
            
            # Formatting recommendation column
            rec_val = str(rec_cell.value)
            if rec_val == "Apply":
                rec_cell.fill = fill_apply
                rec_cell.font = font_apply
                score_cell.fill = fill_apply
                score_cell.font = font_apply
            elif rec_val == "Customize Resume":
                rec_cell.fill = fill_customize
                rec_cell.font = font_customize
                score_cell.fill = fill_customize
                score_cell.font = font_customize
            elif rec_val == "Skip":
                rec_cell.fill = fill_skip
                rec_cell.font = font_skip
                score_cell.fill = fill_skip
                score_cell.font = font_skip
                
            # Apply common formatting to all cells in the row
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border_thin
                
                # Alignments
                if col_num in [1, 2, 10]:  # Score, Recommendation, Posted Date
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Autofit column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Don't let huge URL or descriptions blow out column width
                val = str(cell.value or '')
                if len(val) > 40:
                    val = val[:40] + "..."
                if len(val) > max_len:
                    max_len = len(val)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    return output_path

def _parse_json_field(field: Any) -> Any:
    """Parses text fields stored as JSON in sqlite."""
    if not field:
        return []
    if isinstance(field, str):
        try:
            return json.loads(field)
        except Exception:
            return [field]
    return field
